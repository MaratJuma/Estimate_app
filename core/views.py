from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Q, F, Count
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.http import HttpResponseForbidden, HttpResponse
from django.shortcuts import redirect
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from decimal import Decimal
from django.utils.text import slugify

from .permissions import (
        can_view_estimates,
        can_edit_estimates,
        can_approve_estimates,
        can_manage_services,
        can_manage_contractors,
        deny_access,
        is_admin,
        is_production_manager,
        is_sales_manager,
)
from .models import Service, Estimate, EstimateDay, EstimateItem, Contractor
from .forms import (
        EstimateForm, 
        EstimateItemCreateForm,
        EstimateItemQtyForm, 
        EstimateItemUpdateForm, 
        ContractorForm, 
        # ServiceForm, 
        ServiceCreateForm,
        ServiceForContractorCreateForm,
        ServiceUpdateForm, 
        EstimateDayCreateForm, 
        EstimateDayUpdateForm
)

@login_required
def home(request):
    return render(request, 'core/home.html')

@login_required
def service_list(request):
    query = request.GET.get('q', '').strip()
    selected_category = request.GET.get('category', '').strip()

    services = Service.objects.select_related('contractor').order_by('-id')

    if query:
        services = services.filter(name__icontains=query)

    if selected_category:
        services = services.filter(contractor__category=selected_category)

    categories = (
        Contractor.objects.exclude(category='')
        .values_list('category', flat=True)
        .distinct()
        .order_by('category')
    )

    return render(request, 'core/service_list.html', {
        'services': services,
        'query': query,
        'categories': categories,
        'selected_category': selected_category,
    })

@login_required
def service_detail(request, service_id):

    if not can_manage_services(request.user) and not is_sales_manager(request.user):
        return deny_access(request, 'У вас нет прав на просмотр услуги.')

    service = get_object_or_404(Service.objects.select_related('contractor'), id=service_id)
    next_url = request.GET.get('next')

    return render(request, 'core/service_detail.html', {
        'service': service,
        'next_url': next_url,
    })

@login_required
def service_update(request, service_id):

    if not can_manage_services(request.user):
        return deny_access(request, 'У вас нет прав на управление услугами.')
    
    service = get_object_or_404(Service, id=service_id)
    next_url = request.GET.get('next') or request.POST.get('next')

    old_cost_price = service.cost_price

    if request.method == 'POST':
        form = ServiceUpdateForm(request.POST, instance=service)
        if form.is_valid():
            service = form.save()

            if service.cost_price != old_cost_price:
                estimate_items = list(
                    EstimateItem.objects.filter(
                        service=service,
                        estimate_day__estimate__is_approved=False
                    )
                )

                for item in estimate_items:
                    current_cost_price = item.service.cost_price
                    item.cost_price = service.cost_price
                    item.total_cost = item.qty * service.cost_price

                EstimateItem.objects.bulk_update(
                    estimate_items,
                    ['cost_price', 'total_cost']
                )
            messages.success(request, 'Услуга обновлена.')
            return redirect(next_url or 'service_list')
    else:
        form = ServiceUpdateForm(instance=service)

    return render(request, 'core/service_form.html', {
        'form': form,
        'is_edit': True,
        'service': service,
        'next_url': next_url,
    })

@login_required
def service_delete(request, service_id):

    if not can_manage_services(request.user):
        return deny_access(request, 'У вас нет прав на управление услугами.')
    
    service = get_object_or_404(Service, id=service_id)
    next_url = request.GET.get('next') or request.POST.get('next')

    is_used = EstimateItem.objects.filter(service=service).exists()

    if is_used:
        messages.error(
            request,
            f'Услугу "{service.name}" нельзя удалить, потому что она уже используется в сметах.'
        )
        return redirect(next_url or 'service_list')

    if request.method == 'POST':
        service.delete()
        messages.success(request, f'Услуга "{service.name}" удалена.')
        return redirect(next_url or 'service_list')

    return render(request, 'core/service_confirm_delete.html', {
        'service': service,
        'next_url': next_url,
    })

@login_required
def contractor_list(request):

   # if not can_manage_contractors(request.user):
   #     return deny_access(request, 'У вас нет прав на управление поставщиками.')
    
    query = request.GET.get('q', '').strip()
    selected_category = request.GET.get('category', '').strip()

    # contractors = Contractor.objects.all().order_by('-id')

    contractors = (
        Contractor.objects
        .annotate(services_count=Count('services'))
        .order_by('-id')
    )

    if query:
        contractors = contractors.filter(name__icontains=query)

    if selected_category:
        contractors = contractors.filter(category=selected_category)

    categories = (
        Contractor.objects.exclude(category='')
        .values_list('category', flat=True)
        .distinct()
        .order_by('category')
    )

    return render(request, 'core/contractor_list.html', {
        'contractors': contractors,
        'query': query,
        'categories': categories,
        'selected_category': selected_category,
    })

@login_required
def contractor_update(request, contractor_id):

    if not can_manage_contractors(request.user):
        return deny_access(request, 'У вас нет прав на управление поставщиками.')
    
    contractor = get_object_or_404(Contractor, id=contractor_id)

    if request.method == 'POST':
        form = ContractorForm(request.POST, instance=contractor)
        if form.is_valid():
            form.save()
            return redirect('contractor_detail', contractor_id=contractor.id)
    else:
        form = ContractorForm(instance=contractor)

    return render(request, 'core/contractor_form.html', {
        'form': form,
        'is_edit': True,
        'contractor': contractor,
    })

@login_required
def contractor_detail(request, contractor_id):

    # if not can_manage_contractors(request.user):
    #     return deny_access(request, 'У вас нет прав на просмотр поставщиков.')

    contractor = get_object_or_404(Contractor, id=contractor_id)
    services = contractor.services.all().order_by('name', 'id')
    next_url = request.GET.get('next')

    return render(request, 'core/contractor_detail.html', {
        'contractor': contractor,
        'services': services,
        'next_url': next_url,
    })

@login_required
def estimate_list(request):

    if not can_view_estimates(request.user):
        return deny_access(request, 'У вас нет прав на просмотр смет.')
    
    query = request.GET.get('q', '').strip()

    estimates = (
        Estimate.objects.all()
        .prefetch_related('days__items__service')
        .order_by('-created_at')
    )

    if query:
        estimates = estimates.filter(
            Q(client_name__icontains=query) |
            Q(manager_name__icontains=query) |
            Q(comment__icontains=query)
        )

    for estimate in estimates:
        total_cost = 0
        total_client = 0

        for day in estimate.days.all():
            for item in day.items.all():
                if estimate.is_approved:
                    total_cost += item.total_cost
                    total_client += item.total_client
                else:
                    total_cost += item.qty * item.service.cost_price
                    total_client += item.qty * item.client_price

        estimate.days_count = estimate.days.count()
        estimate.total_cost_sum = total_cost
        estimate.total_client_sum = total_client
        estimate.margin_sum = total_client - total_cost

    return render(request, 'core/estimate_list.html', {
        'estimates': estimates,
        'query': query,
    })

@login_required
def estimate_detail(request, estimate_id):

    if not can_view_estimates(request.user):
        return deny_access(request, 'У вас нет прав на просмотр смет.')

    estimate = get_object_or_404(Estimate, id=estimate_id)
    days = estimate.days.all().order_by('day_number').prefetch_related('items__service')

    total_cost = 0
    total_client = 0

    for day in days:
        day_cost = 0
        day_client = 0

        for item in day.items.all():
            if estimate.is_approved:
                item.display_cost_price = item.cost_price
                item.display_total_cost = item.total_cost
            else:
                current_cost_price = item.service.cost_price
                item.display_cost_price = current_cost_price
                item.display_total_cost = item.qty * current_cost_price

            item.display_client_price = item.client_price
            item.display_total_client = item.qty * item.client_price

            day_cost += item.display_total_cost
            day_client += item.display_total_client

        day.total_cost_sum = day_cost
        day.total_client_sum = day_client

        total_cost += day_cost
        total_client += day_client

    margin = total_client - total_cost
    margin_percent = 0

    if total_client != 0:
        margin_percent = (margin / total_client) * 100

    return render(request, 'core/estimate_detail.html', {
        'estimate': estimate,
        'days': days,
        'total_cost': total_cost,
        'total_client': total_client,
        'margin': margin,
        'margin_percent': margin_percent,
    })

@login_required
def estimate_create(request):

    if not can_edit_estimates(request.user):
        return deny_access(request, 'У вас нет прав на создание смет.')

    if request.method == 'POST':
        form = EstimateForm(request.POST)
        if form.is_valid():
            estimate = form.save()
            EstimateDay.objects.create(estimate=estimate, day_number=1)
            return redirect('estimate_detail', estimate_id=estimate.id)
    else:
        form = EstimateForm()

    return render(request, 'core/estimate_form.html', {
        'form': form,
        'is_edit': False,
        'title': 'Создание сметы',
    })

@login_required
def estimate_item_create(request, day_id):

    if not can_edit_estimates(request.user):
        return deny_access(request, 'У вас нет прав на редактирование смет.')

    day = get_object_or_404(EstimateDay, id=day_id)
    estimate = day.estimate

    if estimate.is_approved:
        messages.error(request, f'Смета #{estimate.id} утверждена. Добавление позиций запрещено.')
        return redirect('estimate_detail', estimate_id=estimate.id)

    query = request.GET.get('q', '').strip()
    selected_category = request.GET.get('category', '').strip()
    selected_contractor = request.GET.get('contractor', '').strip()

    base_services = Service.objects.select_related('contractor').filter(is_active=True)

    services = base_services

    if query:
        services = services.filter(name__icontains=query)

    if selected_category:
        services = services.filter(contractor__category=selected_category)

    if selected_contractor:
        services = services.filter(contractor_id=selected_contractor)

    services = services.order_by('name', 'id')

    categories_qs = base_services

    if query:
        categories_qs = categories_qs.filter(name__icontains=query)

    if selected_contractor:
        categories_qs = categories_qs.filter(contractor_id=selected_contractor)

    categories = (
        categories_qs
        .exclude(contractor__category='')
        .values_list('contractor__category', flat=True)
        .distinct()
        .order_by('contractor__category')
    )

    contractors_qs = base_services

    if query:
        contractors_qs = contractors_qs.filter(name__icontains=query)

    if selected_category:
        contractors_qs = contractors_qs.filter(contractor__category=selected_category)

    contractor_ids = contractors_qs.values_list('contractor_id', flat=True).distinct()

    contractors = Contractor.objects.filter(id__in=contractor_ids).order_by('name')

    selected_contractor_obj = None
    if selected_contractor:
        selected_contractor_obj = Contractor.objects.filter(id=selected_contractor).first()

    return render(request, 'core/estimate_item_form.html', {
        'estimate': estimate,
        'day': day,
        'services': services,
        'query': query,
        'categories': categories,
        'selected_category': selected_category,
        'contractors': contractors,
        'selected_contractor': selected_contractor,
        'selected_contractor_obj': selected_contractor_obj,
        'current_full_path': request.get_full_path(),
    })


@login_required
def estimate_item_create_for_service(request, day_id, service_id):

    if not can_edit_estimates(request.user):
        return deny_access(request, 'У вас нет прав на редактирование смет.')

    day = get_object_or_404(EstimateDay, id=day_id)
    estimate = day.estimate
    service = get_object_or_404(Service.objects.select_related('contractor'), id=service_id, is_active=True)

    if estimate.is_approved:
        messages.error(request, f'Смета #{estimate.id} утверждена. Добавление позиций запрещено.')
        return redirect('estimate_detail', estimate_id=estimate.id)

    if request.method == 'POST':
        form = EstimateItemQtyForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            qty = item.qty

            item.estimate_day = day
            item.service = service
            item.cost_price = service.cost_price
            item.client_price = service.client_price
            item.total_cost = qty * service.cost_price
            item.total_client = qty * service.client_price
            item.save()

            messages.success(request, f'Услуга "{service.name}" добавлена в смету.')
            return redirect('estimate_detail', estimate_id=estimate.id)
    else:
        form = EstimateItemQtyForm()

    return render(request, 'core/estimate_item_add_selected_service.html', {
        'form': form,
        'estimate': estimate,
        'day': day,
        'service': service,
    })

@login_required
def estimate_item_delete(request, item_id):

    if not can_edit_estimates(request.user):
        return deny_access(request, 'У вас нет прав на редактирование смет.')
    
    item = get_object_or_404(EstimateItem, id=item_id)
    estimate_id = item.estimate_day.estimate.id
    estimate = item.estimate_day.estimate

    if estimate.is_approved:
        messages.error(request, f'Смета #{estimate.id} утверждена. Добавление позиций запрещено.')
        return redirect('estimate_detail', estimate_id=estimate.id)

    if request.method == 'POST':
        item.delete()
        return redirect('estimate_detail', estimate_id=estimate_id)

    return render(request, 'core/estimate_item_confirm_delete.html', {
        'item': item,
    })

@login_required
def estimate_item_update(request, item_id):

    if not can_edit_estimates(request.user):
        return deny_access(request, 'У вас нет прав на редактирование смет.')
    
    item = get_object_or_404(EstimateItem, id=item_id)
    estimate = item.estimate_day.estimate

    if estimate.is_approved:
        messages.error(request, f'Смета #{estimate.id} утверждена. Редактирование позиций запрещено.')
        return redirect('estimate_detail', estimate_id=estimate.id)

    display_cost_price = item.service.cost_price

    if request.method == 'POST':
        form = EstimateItemUpdateForm(request.POST, instance=item)
        if form.is_valid():
            item = form.save(commit=False)

            qty = item.qty
            current_cost_price = item.service.cost_price

            item.cost_price = current_cost_price
            item.total_cost = qty * current_cost_price
            item.total_client = qty * item.client_price

            item.save()

            return redirect('estimate_detail', estimate_id=estimate.id)
    else:
        form = EstimateItemUpdateForm(instance=item)

    return render(request, 'core/estimate_item_edit_form.html', {
        'form': form,
        'estimate': estimate,
        'day': item.estimate_day,
        'item': item,
        'is_edit': True,
        'display_cost_price': display_cost_price,
    })

@login_required
def estimate_update(request, estimate_id):

    if not can_edit_estimates(request.user):
        return deny_access(request, 'У вас нет прав на редактирование смет.')
    
    estimate = get_object_or_404(Estimate, id=estimate_id)

    if estimate.is_approved:
        messages.error(request, f'Смета #{estimate.id} утверждена и недоступна для редактирования.')
        return redirect('estimate_detail', estimate_id=estimate.id)

    if request.method == 'POST':
        form = EstimateForm(request.POST, instance=estimate)
        if form.is_valid():
            form.save()
            return redirect('estimate_detail', estimate_id=estimate.id)
    else:
        form = EstimateForm(instance=estimate)

    return render(request, 'core/estimate_form.html', {
        'form': form,
        'is_edit': True,
        'estimate': estimate,
    })

@login_required
def estimate_day_create(request, estimate_id):

    if not can_edit_estimates(request.user):
        return deny_access(request, 'У вас нет прав на редактирование смет.')
    
    estimate = get_object_or_404(Estimate, id=estimate_id)

    if estimate.is_approved:
        messages.error(request, f'Смета #{estimate.id} утверждена. Добавление дней запрещено.')
        return redirect('estimate_detail', estimate_id=estimate.id)

    if request.method == 'POST':
        form = EstimateDayCreateForm(request.POST)
        if form.is_valid():
            day = form.save(commit=False)

            last_day = estimate.days.order_by('-day_number').first()
            next_day_number = last_day.day_number + 1 if last_day else 1

            day.estimate = estimate
            day.day_number = next_day_number
            day.save()

            return redirect('estimate_detail', estimate_id=estimate.id)
    else:
        form = EstimateDayCreateForm()

    return render(request, 'core/estimate_day_form.html', {
        'form': form,
        'estimate': estimate,
        'is_edit': False,
    })

@login_required
def estimate_day_update(request, day_id):

    if not can_edit_estimates(request.user):
        return deny_access(request, 'У вас нет прав на редактирование смет.')
    
    day = get_object_or_404(EstimateDay, id=day_id)
    estimate = day.estimate

    if estimate.is_approved:
        messages.error(request, f'Смета #{estimate.id} утверждена. Редактирование дней запрещено.')
        return redirect('estimate_detail', estimate_id=estimate.id)

    if request.method == 'POST':
        form = EstimateDayUpdateForm(request.POST, instance=day)
        if form.is_valid():
            form.save()
            return redirect('estimate_detail', estimate_id=estimate.id)
    else:
        form = EstimateDayUpdateForm(instance=day)

    return render(request, 'core/estimate_day_form.html', {
        'form': form,
        'estimate': estimate,
        'day': day,
        'is_edit': True,
    })

@login_required
def estimate_day_delete(request, day_id):

    if not can_edit_estimates(request.user):
        return deny_access(request, 'У вас нет прав на редактирование смет.')
    
    day = get_object_or_404(EstimateDay, id=day_id)
    estimate = day.estimate

    if estimate.is_approved:
        messages.error(request, f'Смета #{estimate.id} утверждена. Удаление дней запрещено.')
        return redirect('estimate_detail', estimate_id=estimate.id)


    if day.items.exists():
        messages.error(
            request,
            f'День {day.day_number} нельзя удалить, потому что в нём есть позиции.'
        )
        return redirect('estimate_detail', estimate_id=estimate.id)

    deleted_day_number = day.day_number

    if request.method == 'POST':
        day.delete()

        EstimateDay.objects.filter(
            estimate=estimate,
            day_number__gt=deleted_day_number
        ).update(day_number=F('day_number') - 1)

        messages.success(request, f'День {deleted_day_number} удалён.')
        return redirect('estimate_detail', estimate_id=estimate.id)

    return render(request, 'core/estimate_day_confirm_delete.html', {
        'day': day,
        'estimate': estimate,
    })

@login_required
def estimate_duplicate(request, estimate_id):
    source_estimate = get_object_or_404(
        Estimate.objects.prefetch_related('days__items'),
        id=estimate_id
    )

    if request.method == 'POST':
        if source_estimate.comment:
            new_comment = f"{source_estimate.comment} (копия сметы #{source_estimate.id})"
        else:
            new_comment = f"Копия сметы #{source_estimate.id}"

        new_estimate = Estimate.objects.create(
            client_name=source_estimate.client_name,
            manager_name=source_estimate.manager_name,
            comment=new_comment,
        )

        day_mapping = {}

        for day in source_estimate.days.all():
            new_day = EstimateDay.objects.create(
                estimate=new_estimate,
                day_number=day.day_number,
            )
            day_mapping[day.id] = new_day

        for day in source_estimate.days.all():
            new_day = day_mapping[day.id]

            for item in day.items.all():
                EstimateItem.objects.create(
                    estimate_day=new_day,
                    service=item.service,
                    qty=item.qty,
                    cost_price=item.cost_price,
                    client_price=item.client_price,
                    total_cost=item.total_cost,
                    total_client=item.total_client,
                )

        return redirect('estimate_detail', estimate_id=new_estimate.id)

    return render(request, 'core/estimate_duplicate_confirm.html', {
        'estimate': source_estimate,
    })

@login_required
def contractor_create(request):

    if not can_manage_contractors(request.user):
        return deny_access(request, 'У вас нет прав на управление поставщиками.')
    
    next_url = request.GET.get('next') or request.POST.get('next')

    if request.method == 'POST':
        form = ContractorForm(request.POST)
        if form.is_valid():
            form.save()

            if next_url:
                return redirect(next_url)

            return redirect('service_list')
    else:
        form = ContractorForm()

    return render(request, 'core/contractor_form.html', {
        'form': form,
        'next_url': next_url,
    })


@login_required
def service_create(request):

    if not can_manage_services(request.user):
        return deny_access(request, 'У вас нет прав на управление услугами.')
    
    if request.method == 'POST':
        query = request.POST.get('q', '').strip()
        selected_category = request.POST.get('category', '').strip()
    else:
        query = request.GET.get('q', '').strip()
        selected_category = request.GET.get('category', '').strip()

    contractors = Contractor.objects.all().order_by('name')

    if query:
        contractors = contractors.filter(name__icontains=query)

    if selected_category:
        contractors = contractors.filter(category=selected_category)

    categories = (
        Contractor.objects.values_list('category', flat=True)
        .distinct()
        .order_by('category')
    )

    if request.method == 'POST':
        form = ServiceCreateForm(request.POST)
        form.fields['contractor'].queryset = contractors

        if form.is_valid():
            form.save()
            messages.success(request, 'Услуга создана.')
            return redirect('service_list')
    else:
        form = ServiceCreateForm()
        form.fields['contractor'].queryset = contractors

    return render(request, 'core/service_form.html', {
        'form': form,
        'is_edit': False,
        'query': query,
        'categories': categories,
        'selected_category': selected_category,
    })

@login_required
def service_create_for_contractor(request, contractor_id):

    if not can_manage_services(request.user):
        return deny_access(request, 'У вас нет прав на управление услугами.')

    contractor = get_object_or_404(Contractor, id=contractor_id)

    if request.method == 'POST':
        form = ServiceForContractorCreateForm(request.POST)
        if form.is_valid():
            service = form.save(commit=False)
            service.contractor = contractor
            service.save()
            messages.success(request, 'Услуга создана.')
            return redirect('contractor_detail', contractor_id=contractor.id)
    else:
        form = ServiceForContractorCreateForm()

    return render(request, 'core/service_form.html', {
        'form': form,
        'is_edit': False,
        'contractor': contractor,
        'is_contractor_context': True,
    })


@login_required
def estimate_print(request, estimate_id):
    estimate = get_object_or_404(Estimate, id=estimate_id)
    all_days = estimate.days.all().order_by('day_number').prefetch_related('items__service')
    days = []

    total_client = 0

    for day in all_days:
        items = list(day.items.all())
        if not items:
            continue

        day.items_list = items

        day_total_client = 0
        for item in items:
            day_total_client += item.total_client

        day.total_client_sum = day_total_client
        total_client += day_total_client
        days.append(day)

    company = {
        'name': 'SakhTravel',
        'tagline': 'Объединяя мечты',
        'phone': '+7 (999) 123-45-67',
        'email': 'info@company.ru',
        'site': 'www.sakhtravel.ru',
        'address': 'Южно-Cахалинск, Россия',
        'manager_title': 'Менеджер проекта',
        'manager_name': estimate.manager_name,
    }

    return render(request, 'core/estimate_print.html', {
        'estimate': estimate,
        'days': days,
        'total_client': total_client,
        'company': company,
        'document_date': timezone.now(),
    })


    fonts_dir = os.path.join(settings.BASE_DIR, 'static', 'fonts')
    regular_path = os.path.join(fonts_dir, 'DejaVuSans.ttf')

    if not os.path.exists(regular_path):
        raise FileNotFoundError(f'Не найден шрифт: {regular_path}')

    pdfmetrics.registerFont(TTFont('DejaVuSans', regular_path))


@login_required
def estimate_approve(request, estimate_id):

    if not can_approve_estimates(request.user):
        return deny_access(request, 'У вас нет прав на утверждение смет.')
    
    estimate = get_object_or_404(Estimate, id=estimate_id)

    if estimate.is_approved:
        messages.info(request, f'Смета #{estimate.id} уже утверждена.')
        return redirect('estimate_detail', estimate_id=estimate.id)

    if request.method == 'POST':
        for day in estimate.days.all().prefetch_related('items__service'):
            for item in day.items.all():
                current_cost_price = item.service.cost_price

                item.cost_price = current_cost_price
                item.total_cost = item.qty * current_cost_price
                item.total_client = item.qty * item.client_price
                item.save()

        estimate.is_approved = True
        estimate.approved_at = timezone.now()
        estimate.save()

        messages.success(request, f'Смета #{estimate.id} утверждена. Редактирование заблокировано.')
        return redirect('estimate_detail', estimate_id=estimate.id)

    return render(request, 'core/estimate_approve_confirm.html', {
        'estimate': estimate,
    })


@login_required
def estimate_excel_export(request, estimate_id):
    if not (is_admin(request.user) or is_production_manager(request.user) or is_sales_manager(request.user)):
        return deny_access(request, 'У вас нет прав на экспорт рабочей сметы.')

    estimate = get_object_or_404(Estimate, id=estimate_id)
    days = estimate.days.all().prefetch_related('items__service__contractor').order_by('day_number')

    wb = Workbook()
    ws = wb.active
    ws.title = f'Смета {estimate.id}'

    # Стили
    #base_font_size = 10
    title_font = Font(bold=True, size=13)
    section_font = Font(bold=True, size=10)
    bold_font = Font(bold=True, size=10)
    base_font = Font(size=10)

    title_fill = PatternFill('solid', fgColor='DCE6F1')
    header_fill = PatternFill('solid', fgColor='B8CCE4')
    subtotal_fill = PatternFill('solid', fgColor='EAF2F8')
    total_fill = PatternFill('solid', fgColor='9BC2E6')
    negative_fill = PatternFill('solid', fgColor='F4CCCC')
    warning_fill = PatternFill('solid', fgColor='FFF2CC')
    positive_fill = PatternFill('solid', fgColor='D9EAD3')

    thin_side = Side(style='thin', color='808080')
    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center', wrap_text=True)

    money_format = '#,##0.00'
    qty_format = '#,##0.00'
    percent_format = '0.00%'

    def get_margin_fill(margin_percent):
        if margin_percent < 0:
            return negative_fill
        elif margin_percent <= Decimal('0.10'):
            return warning_fill
        return positive_fill

    row = 1

    # Заголовок
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    cell = ws.cell(row=row, column=1, value=f'Рабочая смета № {estimate.id}')
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = center
    cell.border = thin_border
    ws.row_dimensions[row].height = 22
    row += 2

    # Инфо-блок
    info_rows = [
        ('Клиент', estimate.client_name),
        ('Менеджер', estimate.manager_name),
        ('Дата', estimate.created_at.strftime('%d.%m.%Y %H:%M')),
        ('Статус', 'Утверждена' if estimate.is_approved else 'Черновик'),
    ]

    if estimate.comment:
        info_rows.append(('Коммент.', estimate.comment))

    for label, value in info_rows:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = section_font
        label_cell.fill = title_fill
        label_cell.border = thin_border
        label_cell.alignment = left

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = base_font
        value_cell.border = thin_border
        value_cell.alignment = left

        for col in range(2, 6):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = left
            ws.cell(row=row, column=col).font = base_font

        row += 1

    row += 1
    table_header_row = row

    headers = [
        'День',
        'Название дня',
        'Услуга',
        'Поставщик',
        'Кол-во',
        'Себестоимость',
        'Сумма себестоимости',
        'Цена клиенту',
        'Сумма клиенту',
        'Маржа',
        'Маржа, %',
    ]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    ws.row_dimensions[row].height = 28
    row += 1

    total_cost_sum = Decimal('0')
    total_client_sum = Decimal('0')
    total_margin_sum = Decimal('0')

    day_summary = []

    for day in days:
        day_cost_sum = Decimal('0')
        day_client_sum = Decimal('0')
        day_margin_sum = Decimal('0')
        
        day_items = list(day.items.all().order_by('id'))

        if not day_items:
            ws.cell(row=row, column=1, value=day.day_number)
            ws.cell(row=row, column=2, value=day.title or '')
            ws.cell(row=row, column=3, value='—')

            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.font = base_font
                cell.alignment = left if col in [2, 3, 4] else center if col == 1 else right

            row += 1
        else:
            first_row_for_day = True

            for item in day_items:
                if estimate.is_approved:
                    qty = item.qty or Decimal('0')
                    cost_price = item.cost_price or Decimal('0')
                    total_cost = item.total_cost or Decimal('0')
                    client_price = item.client_price or Decimal('0')
                    total_client = item.total_client or Decimal('0')
                else:
                    qty = item.qty or Decimal('0')
                    cost_price = item.cost_price or Decimal('0')
                    client_price = item.client_price or Decimal('0')
                    total_cost = qty * cost_price
                    total_client = qty * client_price

                margin = total_client - total_cost
                margin_percent = (margin / total_client) if total_client else Decimal('0')

                ws.cell(row=row, column=1, value=day.day_number if first_row_for_day else '')
                ws.cell(row=row, column=2, value=(day.title or '') if first_row_for_day else '')
                ws.cell(row=row, column=3, value=item.service.name if item.service else '')
                ws.cell(
                    row=row,
                    column=4,
                    value=item.service.contractor.name if item.service and item.service.contractor else ''
                )
                ws.cell(row=row, column=5, value=float(qty))
                ws.cell(row=row, column=6, value=float(cost_price))
                ws.cell(row=row, column=7, value=float(total_cost))
                ws.cell(row=row, column=8, value=float(client_price))
                ws.cell(row=row, column=9, value=float(total_client))
                ws.cell(row=row, column=10, value=float(margin))
                ws.cell(row=row, column=11, value=float(margin_percent))

                for col in range(1, 12):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.font = base_font
                    cell.alignment = left if col in [2, 3, 4] else center if col == 1 else right

                ws.cell(row=row, column=5).number_format = qty_format

                for col in [6, 7, 8, 9, 10]:
                    ws.cell(row=row, column=col).number_format = money_format

                ws.cell(row=row, column=11).number_format = percent_format

                if margin < 0:
                    ws.cell(row=row, column=10).fill = negative_fill
                    ws.cell(row=row, column=11).fill = negative_fill

                day_cost_sum += total_cost
                day_client_sum += total_client
                day_margin_sum += margin

                total_cost_sum += total_cost
                total_client_sum += total_client
                total_margin_sum += margin

                row += 1
                first_row_for_day = False

        day_margin_percent = (day_margin_sum / day_client_sum) if day_client_sum else Decimal('0')
        day_summary.append({
            'day_number': day.day_number,
            'day_title': day.title or '',
            'cost_sum': day_cost_sum,
            'client_sum': day_client_sum,
            'margin_sum': day_margin_sum,
            'margin_percent': day_margin_percent,
        })

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        subtotal_label = ws.cell(row=row, column=1, value=f'Итого за день {day.day_number}')
        subtotal_label.font = bold_font
        subtotal_label.fill = subtotal_fill
        subtotal_label.border = thin_border
        subtotal_label.alignment = right

        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.fill = subtotal_fill
            cell.border = thin_border
            cell.font = base_font

        ws.cell(row=row, column=7, value=float(day_cost_sum))
        ws.cell(row=row, column=8, value='')
        ws.cell(row=row, column=9, value=float(day_client_sum))
        ws.cell(row=row, column=10, value=float(day_margin_sum))
        ws.cell(row=row, column=11, value=float(day_margin_percent))

        for col in [7, 9, 10]:
            cell = ws.cell(row=row, column=col)
            cell.font = bold_font
            cell.number_format = money_format
            cell.alignment = right

        ws.cell(row=row, column=11).font = bold_font
        ws.cell(row=row, column=11).number_format = percent_format
        ws.cell(row=row, column=11).alignment = right

        day_margin_fill = get_margin_fill(day_margin_percent)
        ws.cell(row=row, column=10).fill = day_margin_fill
        ws.cell(row=row, column=11).fill = day_margin_fill

        row += 2

    total_margin_percent = (total_margin_sum / total_client_sum) if total_client_sum else Decimal('0')

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    total_label = ws.cell(row=row, column=1, value='ОБЩИЙ ИТОГ')
    total_label.font = Font(bold=True, size=11)
    total_label.fill = total_fill
    total_label.border = thin_border
    total_label.alignment = right

    for col in range(1, 12):
        cell = ws.cell(row=row, column=col)
        cell.fill = total_fill
        cell.border = thin_border
        cell.font = base_font

    ws.cell(row=row, column=7, value=float(total_cost_sum))
    ws.cell(row=row, column=8, value='')
    ws.cell(row=row, column=9, value=float(total_client_sum))
    ws.cell(row=row, column=10, value=float(total_margin_sum))
    ws.cell(row=row, column=11, value=float(total_margin_percent))

    for col in [7, 9, 10]:
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, size=11)
        cell.number_format = money_format
        cell.alignment = right

    ws.cell(row=row, column=11).font = Font(bold=True, size=11)
    ws.cell(row=row, column=11).number_format = percent_format
    ws.cell(row=row, column=11).alignment = right

    total_margin_fill = get_margin_fill(total_margin_percent)
    ws.cell(row=row, column=10).fill = total_margin_fill
    ws.cell(row=row, column=11).fill = total_margin_fill

    widths = {
        'A': 10,
        'B': 20,
        'C': 32,
        'D': 24,
        'E': 10,
        'F': 16,
        'G': 19,
        'H': 16,
        'I': 18,
        'J': 14,
        'K': 12,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = f'A{table_header_row + 1}'
    ws.auto_filter.ref = f'A{table_header_row}:K{row}'

    for r in range(1, row + 1):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = 18

    # Лист "Сводка"
    summary_ws = wb.create_sheet(title='Сводка')

    summary_ws.merge_cells('A1:F1')
    summary_ws['A1'] = f'Сводка по смете № {estimate.id}'
    summary_ws['A1'].font = title_font
    summary_ws['A1'].fill = title_fill
    summary_ws['A1'].alignment = center
    summary_ws['A1'].border = thin_border

    summary_info = [
        ('Клиент', estimate.client_name),
        ('Менеджер', estimate.manager_name),
        ('Дата', estimate.created_at.strftime('%d.%m.%Y %H:%M')),
        ('Статус', 'Утверждена' if estimate.is_approved else 'Черновик'),
        ('Общая себестоимость', float(total_cost_sum)),
        ('Общая сумма клиенту', float(total_client_sum)),
        ('Общая маржа', float(total_margin_sum)),
        ('Общая маржа, %', float(total_margin_percent)),
    ]

    summary_row = 3
    for label, value in summary_info:
        label_cell = summary_ws.cell(row=summary_row, column=1, value=label)
        label_cell.font = section_font
        label_cell.fill = title_fill
        label_cell.border = thin_border
        label_cell.alignment = left

        value_cell = summary_ws.cell(row=summary_row, column=2, value=value)
        value_cell.font = base_font
        value_cell.border = thin_border
        value_cell.alignment = right if isinstance(value, (int, float)) else left

        if label in ['Общая себестоимость', 'Общая сумма клиенту', 'Общая маржа']:
            value_cell.number_format = money_format

        if label == 'Общая маржа, %':
            value_cell.number_format = percent_format
            value_cell.fill = get_margin_fill(total_margin_percent)

        summary_row += 1

    summary_row += 1

    summary_headers = ['День', 'Название дня', 'Себестоимость', 'Сумма клиенту', 'Маржа', 'Маржа, %']
    summary_table_header_row = summary_row

    for col_num, header in enumerate(summary_headers, start=1):
        cell = summary_ws.cell(row=summary_row, column=col_num, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    summary_row += 1

    for day_data in day_summary:
        summary_ws.cell(row=summary_row, column=1, value=day_data['day_number'])
        summary_ws.cell(row=summary_row, column=2, value=day_data['day_title'])
        summary_ws.cell(row=summary_row, column=3, value=float(day_data['cost_sum']))
        summary_ws.cell(row=summary_row, column=4, value=float(day_data['client_sum']))
        summary_ws.cell(row=summary_row, column=5, value=float(day_data['margin_sum']))
        summary_ws.cell(row=summary_row, column=6, value=float(day_data['margin_percent']))

        for col in range(1, 7):
            cell = summary_ws.cell(row=summary_row, column=col)
            cell.border = thin_border
            cell.font = base_font
            cell.alignment = left if col == 2 else right if col >= 3 else center

        for col in [3, 4, 5]:
            summary_ws.cell(row=summary_row, column=col).number_format = money_format

        summary_ws.cell(row=summary_row, column=6).number_format = percent_format

        summary_day_fill = get_margin_fill(day_data['margin_percent'])
        summary_ws.cell(row=summary_row, column=5).fill = summary_day_fill
        summary_ws.cell(row=summary_row, column=6).fill = summary_day_fill

        summary_row += 1

    summary_widths = {
        'A': 20,
        'B': 24,
        'C': 18,
        'D': 18,
        'E': 16,
        'F': 12,
    }

    for col, width in summary_widths.items():
        summary_ws.column_dimensions[col].width = width

    summary_ws.freeze_panes = f'A{summary_table_header_row + 1}'
    summary_ws.auto_filter.ref = f'A{summary_table_header_row}:F{summary_row - 1}'

    for r in range(1, summary_row + 1):
        if summary_ws.row_dimensions[r].height is None:
            summary_ws.row_dimensions[r].height = 18

    # client_slug = slugify(estimate.client_name or 'client') or 'client'
    estimate_date = estimate.created_at.strftime('%Y-%m-%d')
    filename = f'Smeta_{estimate.id}_{estimate_date}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response