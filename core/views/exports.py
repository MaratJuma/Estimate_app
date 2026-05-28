from decimal import Decimal

from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..models import Estimate
from ..permissions import (
    can_view_estimates,
    deny_access,
    is_admin,
    is_production_manager,
    is_sales_manager,
)

def estimate_excel_export(request, estimate_id):
    if not (is_admin(request.user) or is_production_manager(request.user) or is_sales_manager(request.user)):
        return deny_access(request, 'У вас нет прав на экспорт рабочей сметы.')

    estimate = get_object_or_404(Estimate, id=estimate_id)
    days = estimate.days.all().prefetch_related('items__service__contractor').order_by('day_number')

    wb = Workbook()
    ws = wb.active
    ws.title = f'Смета {estimate.id}'

    # Стили
    #base_font_size = 10
    title_font = Font(bold=True, size=13)
    section_font = Font(bold=True, size=10)
    bold_font = Font(bold=True, size=10)
    base_font = Font(size=10)

    title_fill = PatternFill('solid', fgColor='DCE6F1')
    header_fill = PatternFill('solid', fgColor='B8CCE4')
    subtotal_fill = PatternFill('solid', fgColor='EAF2F8')
    total_fill = PatternFill('solid', fgColor='9BC2E6')
    negative_fill = PatternFill('solid', fgColor='F4CCCC')
    warning_fill = PatternFill('solid', fgColor='FFF2CC')
    positive_fill = PatternFill('solid', fgColor='D9EAD3')

    thin_side = Side(style='thin', color='808080')
    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center', wrap_text=True)

    money_format = '#,##0.00'
    qty_format = '#,##0.00'
    percent_format = '0.00%'

    def get_margin_fill(margin_percent):
        if margin_percent < 0:
            return negative_fill
        elif margin_percent <= Decimal('0.10'):
            return warning_fill
        return positive_fill

    row = 1

    # Заголовок
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    cell = ws.cell(row=row, column=1, value=f'Рабочая смета № {estimate.id}')
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = center
    cell.border = thin_border
    ws.row_dimensions[row].height = 22
    row += 2

    # Инфо-блок
    info_rows = [
        ('Клиент', estimate.client_name),
        ('Менеджер', estimate.manager_name),
        ('Дата', estimate.created_at.strftime('%d.%m.%Y %H:%M')),
        ('Статус', 'Утверждена' if estimate.is_approved else 'Черновик'),
    ]

    if estimate.comment:
        info_rows.append(('Коммент.', estimate.comment))

    for label, value in info_rows:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = section_font
        label_cell.fill = title_fill
        label_cell.border = thin_border
        label_cell.alignment = left

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = base_font
        value_cell.border = thin_border
        value_cell.alignment = left

        for col in range(2, 6):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = left
            ws.cell(row=row, column=col).font = base_font

        row += 1

    row += 1
    table_header_row = row

    headers = [
        'День',
        'Название дня',
        'Услуга',
        'Поставщик',
        'Кол-во',
        'Себестоимость',
        'Сумма себестоимости',
        'Цена клиенту',
        'Сумма клиенту',
        'Маржа',
        'Маржа, %',
    ]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    ws.row_dimensions[row].height = 28
    row += 1

    total_cost_sum = Decimal('0')
    total_client_sum = Decimal('0')
    total_margin_sum = Decimal('0')

    day_summary = []

    for day in days:
        day_cost_sum = Decimal('0')
        day_client_sum = Decimal('0')
        day_margin_sum = Decimal('0')
        
        day_items = list(day.items.all().order_by('id'))

        if not day_items:
            ws.cell(row=row, column=1, value=day.day_number)
            ws.cell(row=row, column=2, value=day.title or '')
            ws.cell(row=row, column=3, value='—')

            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.font = base_font
                cell.alignment = left if col in [2, 3, 4] else center if col == 1 else right

            row += 1
        else:
            first_row_for_day = True

            for item in day_items:
                qty = item.qty or Decimal('0')
                cost_price = item.cost_price or Decimal('0')
                client_price = item.client_price or Decimal('0')
                total_cost = item.total_cost or Decimal('0')
                total_client = item.total_client or Decimal('0')

                margin = total_client - total_cost
                margin_percent = (margin / total_client) if total_client else Decimal('0')

                ws.cell(row=row, column=1, value=day.day_number if first_row_for_day else '')
                ws.cell(row=row, column=2, value=(day.title or '') if first_row_for_day else '')
                ws.cell(row=row, column=3, value=item.service.name if item.service else '')
                ws.cell(
                    row=row,
                    column=4,
                    value=item.service.contractor.name if item.service and item.service.contractor else ''
                )
                ws.cell(row=row, column=5, value=float(qty))
                ws.cell(row=row, column=6, value=float(cost_price))
                ws.cell(row=row, column=7, value=float(total_cost))
                ws.cell(row=row, column=8, value=float(client_price))
                ws.cell(row=row, column=9, value=float(total_client))
                ws.cell(row=row, column=10, value=float(margin))
                ws.cell(row=row, column=11, value=float(margin_percent))

                for col in range(1, 12):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.font = base_font
                    cell.alignment = left if col in [2, 3, 4] else center if col == 1 else right

                ws.cell(row=row, column=5).number_format = qty_format

                for col in [6, 7, 8, 9, 10]:
                    ws.cell(row=row, column=col).number_format = money_format

                ws.cell(row=row, column=11).number_format = percent_format

                if margin < 0:
                    ws.cell(row=row, column=10).fill = negative_fill
                    ws.cell(row=row, column=11).fill = negative_fill

                day_cost_sum += total_cost
                day_client_sum += total_client
                day_margin_sum += margin

                total_cost_sum += total_cost
                total_client_sum += total_client
                total_margin_sum += margin

                row += 1
                first_row_for_day = False

        day_margin_percent = (day_margin_sum / day_client_sum) if day_client_sum else Decimal('0')
        day_summary.append({
            'day_number': day.day_number,
            'day_title': day.title or '',
            'cost_sum': day_cost_sum,
            'client_sum': day_client_sum,
            'margin_sum': day_margin_sum,
            'margin_percent': day_margin_percent,
        })

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        subtotal_label = ws.cell(row=row, column=1, value=f'Итого за день {day.day_number}')
        subtotal_label.font = bold_font
        subtotal_label.fill = subtotal_fill
        subtotal_label.border = thin_border
        subtotal_label.alignment = right

        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.fill = subtotal_fill
            cell.border = thin_border
            cell.font = base_font

        ws.cell(row=row, column=7, value=float(day_cost_sum))
        ws.cell(row=row, column=8, value='')
        ws.cell(row=row, column=9, value=float(day_client_sum))
        ws.cell(row=row, column=10, value=float(day_margin_sum))
        ws.cell(row=row, column=11, value=float(day_margin_percent))

        for col in [7, 9, 10]:
            cell = ws.cell(row=row, column=col)
            cell.font = bold_font
            cell.number_format = money_format
            cell.alignment = right

        ws.cell(row=row, column=11).font = bold_font
        ws.cell(row=row, column=11).number_format = percent_format
        ws.cell(row=row, column=11).alignment = right

        day_margin_fill = get_margin_fill(day_margin_percent)
        ws.cell(row=row, column=10).fill = day_margin_fill
        ws.cell(row=row, column=11).fill = day_margin_fill

        row += 2

    total_margin_percent = (total_margin_sum / total_client_sum) if total_client_sum else Decimal('0')

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    total_label = ws.cell(row=row, column=1, value='ОБЩИЙ ИТОГ')
    total_label.font = Font(bold=True, size=11)
    total_label.fill = total_fill
    total_label.border = thin_border
    total_label.alignment = right

    for col in range(1, 12):
        cell = ws.cell(row=row, column=col)
        cell.fill = total_fill
        cell.border = thin_border
        cell.font = base_font

    ws.cell(row=row, column=7, value=float(total_cost_sum))
    ws.cell(row=row, column=8, value='')
    ws.cell(row=row, column=9, value=float(total_client_sum))
    ws.cell(row=row, column=10, value=float(total_margin_sum))
    ws.cell(row=row, column=11, value=float(total_margin_percent))

    for col in [7, 9, 10]:
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, size=11)
        cell.number_format = money_format
        cell.alignment = right

    ws.cell(row=row, column=11).font = Font(bold=True, size=11)
    ws.cell(row=row, column=11).number_format = percent_format
    ws.cell(row=row, column=11).alignment = right

    total_margin_fill = get_margin_fill(total_margin_percent)
    ws.cell(row=row, column=10).fill = total_margin_fill
    ws.cell(row=row, column=11).fill = total_margin_fill

    widths = {
        'A': 10,
        'B': 20,
        'C': 32,
        'D': 24,
        'E': 10,
        'F': 16,
        'G': 19,
        'H': 16,
        'I': 18,
        'J': 14,
        'K': 12,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = f'A{table_header_row + 1}'
    ws.auto_filter.ref = f'A{table_header_row}:K{row}'

    for r in range(1, row + 1):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = 18

    # Лист "Сводка"
    summary_ws = wb.create_sheet(title='Сводка')

    summary_ws.merge_cells('A1:F1')
    summary_ws['A1'] = f'Сводка по смете № {estimate.id}'
    summary_ws['A1'].font = title_font
    summary_ws['A1'].fill = title_fill
    summary_ws['A1'].alignment = center
    summary_ws['A1'].border = thin_border

    summary_info = [
        ('Клиент', estimate.client_name),
        ('Менеджер', estimate.manager_name),
        ('Дата', estimate.created_at.strftime('%d.%m.%Y %H:%M')),
        ('Статус', 'Утверждена' if estimate.is_approved else 'Черновик'),
        ('Общая себестоимость', float(total_cost_sum)),
        ('Общая сумма клиенту', float(total_client_sum)),
        ('Общая маржа', float(total_margin_sum)),
        ('Общая маржа, %', float(total_margin_percent)),
    ]

    summary_row = 3
    for label, value in summary_info:
        label_cell = summary_ws.cell(row=summary_row, column=1, value=label)
        label_cell.font = section_font
        label_cell.fill = title_fill
        label_cell.border = thin_border
        label_cell.alignment = left

        value_cell = summary_ws.cell(row=summary_row, column=2, value=value)
        value_cell.font = base_font
        value_cell.border = thin_border
        value_cell.alignment = right if isinstance(value, (int, float)) else left

        if label in ['Общая себестоимость', 'Общая сумма клиенту', 'Общая маржа']:
            value_cell.number_format = money_format

        if label == 'Общая маржа, %':
            value_cell.number_format = percent_format
            value_cell.fill = get_margin_fill(total_margin_percent)

        summary_row += 1

    summary_row += 1

    summary_headers = ['День', 'Название дня', 'Себестоимость', 'Сумма клиенту', 'Маржа', 'Маржа, %']
    summary_table_header_row = summary_row

    for col_num, header in enumerate(summary_headers, start=1):
        cell = summary_ws.cell(row=summary_row, column=col_num, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    summary_row += 1

    for day_data in day_summary:
        summary_ws.cell(row=summary_row, column=1, value=day_data['day_number'])
        summary_ws.cell(row=summary_row, column=2, value=day_data['day_title'])
        summary_ws.cell(row=summary_row, column=3, value=float(day_data['cost_sum']))
        summary_ws.cell(row=summary_row, column=4, value=float(day_data['client_sum']))
        summary_ws.cell(row=summary_row, column=5, value=float(day_data['margin_sum']))
        summary_ws.cell(row=summary_row, column=6, value=float(day_data['margin_percent']))

        for col in range(1, 7):
            cell = summary_ws.cell(row=summary_row, column=col)
            cell.border = thin_border
            cell.font = base_font
            cell.alignment = left if col == 2 else right if col >= 3 else center

        for col in [3, 4, 5]:
            summary_ws.cell(row=summary_row, column=col).number_format = money_format

        summary_ws.cell(row=summary_row, column=6).number_format = percent_format

        summary_day_fill = get_margin_fill(day_data['margin_percent'])
        summary_ws.cell(row=summary_row, column=5).fill = summary_day_fill
        summary_ws.cell(row=summary_row, column=6).fill = summary_day_fill

        summary_row += 1

    summary_widths = {
        'A': 20,
        'B': 24,
        'C': 18,
        'D': 18,
        'E': 16,
        'F': 12,
    }

    for col, width in summary_widths.items():
        summary_ws.column_dimensions[col].width = width

    summary_ws.freeze_panes = f'A{summary_table_header_row + 1}'
    summary_ws.auto_filter.ref = f'A{summary_table_header_row}:F{summary_row - 1}'

    for r in range(1, summary_row + 1):
        if summary_ws.row_dimensions[r].height is None:
            summary_ws.row_dimensions[r].height = 18

    # client_slug = slugify(estimate.client_name or 'client') or 'client'
    estimate_date = estimate.created_at.strftime('%Y-%m-%d')
    filename = f'Smeta_{estimate.id}_{estimate_date}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

def estimate_print(request, estimate_id):
    if not can_view_estimates(request.user):
        return deny_access(request, 'У вас нет прав на просмотр смет.')

    estimate = get_object_or_404(Estimate, id=estimate_id)
    all_days = estimate.days.all().order_by('day_number').prefetch_related('items__service')
    days = []

    total_client = Decimal('0.00')

    for day in all_days:
        items = list(day.items.all())
        if not items:
            continue

        day.items_list = items

        day_total_client = Decimal('0.00')
        for item in items:
            day_total_client += item.total_client

        day.total_client_sum = day_total_client
        total_client += day_total_client
        days.append(day)

    company = {
        'name': 'SAKHTRAVEL',
        'tagline': 'Объединяя мечты',
        'phone': '+7 (934) 477-30-08',
        'email': 'go@sakhtravel.com',
        'site': 'www.sakhtravel.com',
        'address': 'г. Южно-Сахалинск, Есенина 1',
        'manager_title': 'Менеджер проекта',
        'manager_name': estimate.manager_name,
    }

    return render(request, 'core/estimate_print.html', {
        'estimate': estimate,
        'days': days,
        'total_client': total_client,
        'company': company,
        'document_date': timezone.now(),
    })