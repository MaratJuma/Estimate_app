from decimal import Decimal, InvalidOperation

from django.db import IntegrityError, transaction
from openpyxl import load_workbook

from core.models import Contractor, Service, ServiceCategory


EXPECTED_HEADERS = [
    'Подрядчик',
    'Комментарий к подрядчику',
    'Телефон',
    'E-mail',
    'Категория услуги',
    'Услуга',
    'Комментарий к услуге',
    'Себестоимость',
    'Цена',
    'Медиа',
]

PREVIEW_LIMIT = 20


def normalize_string(value):
    if value is None:
        return ''
    return str(value).strip()


def normalize_header(value):
    return normalize_string(value).lower()


def normalize_entity_name(value):
    return normalize_string(value)


def make_cache_key(value):
    return normalize_entity_name(value).lower()


def parse_decimal(value):
    if value is None or str(value).strip() == '':
        raise ValueError('Пустое числовое значение')

    raw = str(value).strip().replace(' ', '').replace(',', '.')
    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError):
        raise ValueError(f'Некорректное числовое значение: {value}')


def validate_headers(actual_headers):
    normalized_actual = [normalize_header(h) for h in actual_headers]
    normalized_expected = [normalize_header(h) for h in EXPECTED_HEADERS]

    if normalized_actual != normalized_expected:
        raise ValueError(
            'Неверная структура файла. '
            'Ожидаются столбцы: ' + ' | '.join(EXPECTED_HEADERS)
        )


def build_empty_result(*, dry_run):
    return {
        'dry_run': dry_run,
        'rows_processed': 0,
        'rows_skipped': 0,
        'categories_created': 0,
        'categories_found': 0,
        'contractors_created': 0,
        'contractors_updated': 0,
        'contractors_found': 0,
        'services_created': 0,
        'services_updated': 0,
        'errors': [],
        'preview_rows': [],
    }


def contractor_needs_update(contractor, *, notes, phone, email):
    if notes and contractor.notes != notes:
        return True
    if phone and contractor.phone != phone:
        return True
    if email and contractor.email != email:
        return True
    return False


def get_or_create_category(name, stats, *, dry_run=False, category_cache=None):
    category_name = normalize_entity_name(name)
    cache_key = make_cache_key(category_name)

    if not category_name:
        raise ValueError('Пустое имя категории.')

    if category_cache is not None and cache_key in category_cache:
        cached = category_cache[cache_key]
        if cached['exists_in_db']:
            stats['categories_found'] += 1
            return cached['object'], 'found'
        stats['categories_found'] += 1
        return None, 'found'

    category = ServiceCategory.objects.filter(name__iexact=category_name).first()

    if category:
        stats['categories_found'] += 1
        if category_cache is not None:
            category_cache[cache_key] = {
                'object': category,
                'exists_in_db': True,
            }
        return category, 'found'

    stats['categories_created'] += 1

    if dry_run:
        if category_cache is not None:
            category_cache[cache_key] = {
                'object': None,
                'exists_in_db': False,
            }
        return None, 'create'

    try:
        category = ServiceCategory.objects.create(name=category_name)
    except IntegrityError:
        category = ServiceCategory.objects.get(name__iexact=category_name)

    if category_cache is not None:
        category_cache[cache_key] = {
            'object': category,
            'exists_in_db': True,
        }

    return category, 'create'


def get_or_create_or_update_contractor(
    name,
    notes,
    phone,
    email,
    stats,
    *,
    dry_run=False,
    contractor_cache=None,
):
    contractor_name = normalize_entity_name(name)
    cache_key = make_cache_key(contractor_name)

    if not contractor_name:
        raise ValueError('Пустое имя подрядчика.')

    notes = normalize_string(notes)
    phone = normalize_string(phone)
    email = normalize_string(email)

    if contractor_cache is not None and cache_key in contractor_cache:
        cached = contractor_cache[cache_key]

        if cached['exists_in_db']:
            contractor = cached['object']

            if contractor is not None and contractor_needs_update(
                contractor,
                notes=notes,
                phone=phone,
                email=email,
            ):
                stats['contractors_updated'] += 1

                if not dry_run:
                    if notes:
                        contractor.notes = notes
                    if phone:
                        contractor.phone = phone
                    if email:
                        contractor.email = email
                    contractor.save()

                return contractor, 'update'

            stats['contractors_found'] += 1
            return contractor, 'found'

        preview_notes = cached.get('notes', '')
        preview_phone = cached.get('phone', '')
        preview_email = cached.get('email', '')

        would_update = (
            (notes and preview_notes != notes) or
            (phone and preview_phone != phone) or
            (email and preview_email != email)
        )

        if would_update:
            stats['contractors_updated'] += 1
            cached['notes'] = notes or preview_notes
            cached['phone'] = phone or preview_phone
            cached['email'] = email or preview_email
            return None, 'update'

        stats['contractors_found'] += 1
        return None, 'found'

    contractor = Contractor.objects.filter(name__iexact=contractor_name).first()

    if contractor:
        if contractor_cache is not None:
            contractor_cache[cache_key] = {
                'object': contractor,
                'exists_in_db': True,
            }

        if contractor_needs_update(contractor, notes=notes, phone=phone, email=email):
            stats['contractors_updated'] += 1

            if not dry_run:
                if notes:
                    contractor.notes = notes
                if phone:
                    contractor.phone = phone
                if email:
                    contractor.email = email
                contractor.save()

            return contractor, 'update'

        stats['contractors_found'] += 1
        return contractor, 'found'

    stats['contractors_created'] += 1

    if dry_run:
        if contractor_cache is not None:
            contractor_cache[cache_key] = {
                'object': None,
                'exists_in_db': False,
                'notes': notes,
                'phone': phone,
                'email': email,
            }
        return None, 'create'

    try:
        contractor = Contractor.objects.create(
            name=contractor_name,
            notes=notes,
            phone=phone,
            email=email,
        )
    except IntegrityError:
        contractor = Contractor.objects.get(name__iexact=contractor_name)

        if contractor_needs_update(contractor, notes=notes, phone=phone, email=email):
            if notes:
                contractor.notes = notes
            if phone:
                contractor.phone = phone
            if email:
                contractor.email = email
            contractor.save()
            stats['contractors_updated'] += 1

    if contractor_cache is not None:
        contractor_cache[cache_key] = {
            'object': contractor,
            'exists_in_db': True,
        }

    return contractor, 'create'


def create_or_update_service(
    *,
    contractor,
    category,
    contractor_name,
    category_name,
    name,
    description,
    cost_price,
    client_price,
    image_url,
    stats,
    dry_run=False,
    service_cache=None,
):
    service_name = normalize_string(name)
    description = normalize_string(description)
    image_url = normalize_string(image_url)

    if not service_name:
        raise ValueError('Пустое имя услуги.')

    service_cache_key = (
        make_cache_key(contractor_name),
        make_cache_key(category_name),
        service_name.lower(),
    )

    if service_cache is not None and service_cache_key in service_cache:
        cached = service_cache[service_cache_key]

        if cached['exists_in_db']:
            service = cached['object']

            if not dry_run:
                service.description = description
                service.cost_price = cost_price
                service.client_price = client_price
                service.image_url = image_url
                service.is_active = True
                service.save()

            stats['services_updated'] += 1
            return service, 'update'

        stats['services_updated'] += 1
        return None, 'update'

    if dry_run:
        existing_contractor = Contractor.objects.filter(
            name__iexact=normalize_entity_name(contractor_name)
        ).first()
        existing_category = ServiceCategory.objects.filter(
            name__iexact=normalize_entity_name(category_name)
        ).first()

        existing_service = None
        if existing_contractor and existing_category:
            existing_service = Service.objects.filter(
                contractor=existing_contractor,
                category=existing_category,
                name__iexact=service_name,
            ).first()

        if existing_service:
            stats['services_updated'] += 1
            if service_cache is not None:
                service_cache[service_cache_key] = {
                    'object': existing_service,
                    'exists_in_db': True,
                }
            return None, 'update'

        stats['services_created'] += 1
        if service_cache is not None:
            service_cache[service_cache_key] = {
                'object': None,
                'exists_in_db': False,
            }
        return None, 'create'

    service = Service.objects.filter(
        contractor=contractor,
        category=category,
        name__iexact=service_name,
    ).first()

    if service:
        service.description = description
        service.cost_price = cost_price
        service.client_price = client_price
        service.image_url = image_url
        service.is_active = True
        service.save()

        stats['services_updated'] += 1
        if service_cache is not None:
            service_cache[service_cache_key] = {
                'object': service,
                'exists_in_db': True,
            }
        return service, 'update'

    service = Service.objects.create(
        contractor=contractor,
        category=category,
        name=service_name,
        description=description,
        cost_price=cost_price,
        client_price=client_price,
        image_url=image_url,
        is_active=True,
    )
    stats['services_created'] += 1

    if service_cache is not None:
        service_cache[service_cache_key] = {
            'object': service,
            'exists_in_db': True,
        }

    return service, 'create'


def append_preview_row(
    stats,
    *,
    row_number,
    category_name,
    service_name,
    contractor_name,
    cost_price,
    client_price,
    category_action,
    contractor_action,
    service_action,
):
    if len(stats['preview_rows']) >= PREVIEW_LIMIT:
        return

    stats['preview_rows'].append({
        'row_number': row_number,
        'category_name': category_name,
        'service_name': service_name,
        'contractor_name': contractor_name,
        'cost_price': str(cost_price) if cost_price is not None else '',
        'client_price': str(client_price) if client_price is not None else '',
        'category_action': category_action,
        'contractor_action': contractor_action,
        'service_action': service_action,
    })


def process_import(file_obj, *, dry_run=True):
    workbook = load_workbook(filename=file_obj, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise ValueError('Файл пустой.')

    headers = list(rows[0])
    validate_headers(headers)

    stats = build_empty_result(dry_run=dry_run)

    category_cache = {}
    contractor_cache = {}
    service_cache = {}

    current_contractor_name = ''
    current_contractor_comment = ''
    current_phone = ''
    current_email = ''
    current_category_name = ''

    for index, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            stats['rows_skipped'] += 1
            continue

        try:
            contractor_name_raw = normalize_entity_name(row[0])
            contractor_comment_raw = normalize_string(row[1])
            phone_raw = normalize_string(row[2])
            email_raw = normalize_string(row[3])
            category_name_raw = normalize_entity_name(row[4])
            service_name = normalize_string(row[5])
            service_comment = normalize_string(row[6])
            cost_price_raw = row[7]
            client_price_raw = row[8]
            image_url = normalize_string(row[9])

            if contractor_name_raw:
                current_contractor_name = contractor_name_raw
                current_contractor_comment = contractor_comment_raw
                current_phone = phone_raw
                current_email = email_raw
                current_category_name = ''

            elif contractor_comment_raw or phone_raw or email_raw:
                if not current_contractor_name:
                    raise ValueError(
                        'Указаны данные подрядчика без имени подрядчика и без активного контекста подрядчика.'
                    )

                if contractor_comment_raw:
                    current_contractor_comment = contractor_comment_raw
                if phone_raw:
                    current_phone = phone_raw
                if email_raw:
                    current_email = email_raw

            if not current_contractor_name:
                raise ValueError(
                    'Не заполнено поле "Подрядчик" и отсутствует предыдущий подрядчик для наследования.'
                )

            if category_name_raw:
                current_category_name = category_name_raw

            if service_name and not current_category_name:
                raise ValueError(
                    'Не заполнено поле "Категория услуги" и отсутствует предыдущая категория для наследования.'
                )

            contractor, contractor_action = get_or_create_or_update_contractor(
                current_contractor_name,
                current_contractor_comment,
                current_phone,
                current_email,
                stats,
                dry_run=dry_run,
                contractor_cache=contractor_cache,
            )

            if not service_name:
                stats['rows_processed'] += 1
                continue

            cost_price = parse_decimal(cost_price_raw)
            client_price = parse_decimal(client_price_raw)

            category, category_action = get_or_create_category(
                current_category_name,
                stats,
                dry_run=dry_run,
                category_cache=category_cache,
            )

            _, service_action = create_or_update_service(
                contractor=contractor,
                category=category,
                contractor_name=current_contractor_name,
                category_name=current_category_name,
                name=service_name,
                description=service_comment,
                cost_price=cost_price,
                client_price=client_price,
                image_url=image_url,
                stats=stats,
                dry_run=dry_run,
                service_cache=service_cache,
            )

            append_preview_row(
                stats,
                row_number=index,
                category_name=current_category_name,
                service_name=service_name,
                contractor_name=current_contractor_name,
                cost_price=cost_price,
                client_price=client_price,
                category_action=category_action,
                contractor_action=contractor_action,
                service_action=service_action,
            )

            stats['rows_processed'] += 1

        except Exception as exc:
            stats['errors'].append({
                'row': index,
                'message': str(exc),
            })

    return stats


def preview_services_import(file_obj):
    return process_import(file_obj, dry_run=True)


@transaction.atomic
def import_services_from_excel(file_obj):
    return process_import(file_obj, dry_run=False)